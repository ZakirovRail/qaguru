import pytest
import csv
import glob
import os.path
import pathlib
import zipfile
import shutil
import PyPDF2
import pandas as pd

from io import BytesIO
from openpyxl import Workbook
from reportlab.pdfgen.canvas import Canvas


"""
1. Write a function for preparing pdf, xlsx and csv files in the same directory as executable file
2. Write a function for archiving prepared files and moving archive to the folder "resources"
3. Write a test for reading files in archive without unzipping files 
"""


def prepare_test_files() -> None:
    if not os.path.isdir("test_data"):
        os.mkdir("test_data")
    #  Preparing csv test file
    with open("test_data/test_csv.csv", "w") as test_file:
        csv_writer = csv.writer(test_file, delimiter=";")
        csv_writer.writerow(["Name", "Surname", "Age", "Email", "Registration Date"])
        csv_writer.writerow(["Alena", "Suvorova", "35", "test_email1@gmail.com", "01.01.2023 21:30:00"])
        csv_writer.writerow(["Alex", "McCartny", "45", "test_email2@gmail.com", "02.01.2023 21:30:00"])
        csv_writer.writerow(["Chris", "Martiny", "40", "test_email3@gmail.com", "03.01.2023 21:30:00"])
        csv_writer.writerow(["Ondrej", "Davinchi", "30", "test_email4@gmail.com", "04.01.2023 21:30:00"])
        csv_writer.writerow(["Petra", "Kralovna", "25", "test_email5@gmail.com", "05.01.2023 21:30:00"])

    #  Preparing xlsx test file
    wb = Workbook()
    sheet = wb.active
    a1 = sheet.cell(row=1, column=1)
    a1.value = "Name"
    b1 = sheet.cell(row=1, column=2)
    b1.value = "Surname"
    c1 = sheet.cell(row=1, column=3)
    c1.value = "Age"
    d1 = sheet.cell(row=1, column=4)
    d1.value = "Email"
    e1 = sheet.cell(row=1, column=5)
    e1.value = "Registration Date"

    a2 = sheet.cell(row=2, column=1)
    a2.value = "Alena"
    b2 = sheet.cell(row=2, column=2)
    b2.value = "Suvorova"
    c2 = sheet.cell(row=2, column=3)
    c2.value = "35"
    d2 = sheet.cell(row=2, column=4)
    d2.value = "test_email1@gmail.com"
    e2 = sheet.cell(row=2, column=5)
    e2.value = "01.01.2023 21:30:00"

    a3 = sheet.cell(row=3, column=1)
    a3.value = "Alex"
    b3 = sheet.cell(row=3, column=2)
    b3.value = "McCartny"
    c3 = sheet.cell(row=3, column=3)
    c3.value = "45"
    d3 = sheet.cell(row=3, column=4)
    d3.value = "test_email2@gmail.com"
    e3 = sheet.cell(row=3, column=5)
    e3.value = "02.01.2023 21:30:00"

    wb.save("test_data/test_xlsx.xlsx")

    #  Preparing pdf test file
    canvas = Canvas("test_data/test_pdf.pdf", pagesize=(612.0, 792.0))
    canvas.drawString(1, 780.0, "Name, Surname, Age, Email, Registration Date")
    canvas.drawString(1, 760.0, "Alena, Suvorova, 35, test_email1@gmail.com, 01.01.2023 21:30:00")
    canvas.drawString(1, 740.0, "Alex, McCartny, 45, test_email2@gmail.com, 02.01.2023 21:30:00")
    canvas.drawString(1, 720.0, "Chris, Martiny, 40, test_email3@gmail.com, 03.01.2023 21:30:00")
    canvas.drawString(1, 700.0, "Ondrej, Davinchi, 30, test_email4@gmail.com, 04.01.2023 21:30:00")
    canvas.drawString(1, 680.0, "Petra, Kralovna, 25, test_email5@gmail.com, 05.01.2023 21:30:00")

    canvas.save()


def zip_all_files_in_dir() -> None:
    with zipfile.ZipFile("resources/test_archive.zip", "w", compression=zipfile.ZIP_DEFLATED) as final_archive:
        for file in glob.glob("test_data/test_*"):
            if file != '.DS_Store' or file != 'test_archive.zip':
                add_file = os.path.join("", file)
                print(f"add_file:{add_file}")
                final_archive.write(add_file)
    print("We created  a zip archive")


@pytest.fixture
def archive_move_files():
    print("Generating test files")
    prepare_test_files()
    print("Archiving test files")
    zip_all_files_in_dir()

    yield

    print("\n Now we should remove all test files after running test")
    # If you want manually check that test data and archive were created, disable bellow lines
    if pathlib.Path("resources/test_archive.zip").is_file():
        print("Removing the archive:'test_archive.zip'")
        os.remove("resources/test_archive.zip")
    if pathlib.Path("test_data").is_dir():
        print("Removing the dir:'test_data'")
        shutil.rmtree(pathlib.Path("test_data"))


def test_check_files_content_in_archive(archive_move_files):
    print("Running test for checking content inside the archive")
    with zipfile.ZipFile("resources/test_archive.zip", "r") as zip_:
        for file in zip_.infolist():
            print(f"file is: {os.path.basename(file.filename)}")
            file_extension = (os.path.basename(file.filename)).split(".")[-1]
            print(f"extension of file is: {file_extension}")
            if file_extension == 'csv':
                text = zip_.open(file).read().decode()
                assert "Name" in text
                assert "Surname" in text
                assert "Age" in text
                assert "Email" in text
                assert "Registration Date" in text
            elif file_extension == 'xlsx':
                df = pd.read_excel(zip_.open(file))
                assert "Name" in df
                assert "Surname" in df
                assert "Age" in df
                assert "Email" in df
                assert "Registration Date" in df
            elif file_extension == 'pdf':
                pdf_file = PyPDF2.PdfReader(BytesIO(zip_.read(file)))
                text = pdf_file.pages[0].extract_text()
                assert "Name" in text
                assert "Surname" in text
                assert "Age" in text
                assert "Email" in text
                assert "Registration Date" in text
